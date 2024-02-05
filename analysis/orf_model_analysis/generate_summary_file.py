import argparse
import json
import os
import typing
from collections import defaultdict
from pathlib import Path

import openpyxl


aa_list = ["C", "D", "S", "Q", "M", "N", "P", "K", "_", "T", "F", "A", "G", "I", "L", "H", "R", "W", "V", "E", "Y"]


def accumulate_summary_files(results_directory: str,
                             description_to_gene_file_path: str,
                             gene_to_longest_sequence_file_path: str,
                             optimization_method: str) -> None:
    filename = "summary.xlsx"

    with open(description_to_gene_file_path, "r") as genes_file:
        description_to_gene_mapping = json.load(genes_file)

    with open(gene_to_longest_sequence_file_path, "r") as genes_file:
        gene_to_longest_sequence = json.load(genes_file)

    columns_per_organism_count = 4
    general_columns_count = 4 + len(aa_list)
    for root, dirs, files in os.walk(results_directory):
        for file in files:
            if file == filename:
                sequence_name = Path(root).name.replace("-", "|")
                gene_name = description_to_gene_mapping.get(sequence_name)
                if gene_to_longest_sequence.get(gene_name) != sequence_name:
                    continue

                file_path = os.path.join(root, file)
                try:
                    workbook = openpyxl.load_workbook(filename=file_path)
                    worksheet = workbook.active
                except:
                    print(F"Failed to load: {file_path}")
                    continue

                original_headers = [cell.value for cell in worksheet[1]]
                headers = ["Sequence", "Gene"] + original_headers
                optimization_method_to_scores = {}
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    if row[0] != optimization_method:
                        continue
                    # optimization_method = models.OptimizationMethod(value=row[1])

                    organisms_optimization_index = "_".join(
                        F"{original_headers[i]}_{str(row[i])}" for i in
                        range(1, len(row) - general_columns_count, columns_per_organism_count)
                    )
                    key = "-".join([row[0], organisms_optimization_index])

                    # Needed to remove duplicates from the summary file
                    optimization_method_to_scores[key] = (sequence_name, gene_name) + row

                for index, row in optimization_method_to_scores.items():
                    summary_file_path = F"{index}-summary.xlsx"
                    if not os.path.exists(summary_file_path):
                        summary_workbook = openpyxl.Workbook()
                        summary_worksheet = summary_workbook.active
                        summary_worksheet.append(headers)
                        summary_worksheet.append(row)
                        summary_workbook.save(summary_file_path)
                    else:
                        summary_workbook = openpyxl.load_workbook(summary_file_path)
                        summary_worksheet = summary_workbook.active
                        summary_worksheet.append(row)
                        summary_workbook.save(summary_file_path)


def generate_summary(results_directory: str) -> None:
    filename = "run_summary.json"

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    ordered_organisms = None
    for root, dirs, files in os.walk(results_directory):
        for file in files:
            if file == filename:
                file_path = os.path.join(root, file)
                if ordered_organisms:
                    update_from_summary(file_path=file_path,
                                        worksheet=worksheet,
                                        ordered_organisms=ordered_organisms)
                else:
                    ordered_organisms = update_from_summary(file_path=file_path, worksheet=worksheet)

    workbook.save(os.path.join(results_directory, "summary.xlsx"))


def accumulate_summary_files_to_json(results_directory: str,
                                     description_to_gene_file_path: str,
                                     gene_to_longest_sequence_file_path: str,
                                     optimization_method: str) -> None:
    filename = "summary.xlsx"

    with open(description_to_gene_file_path, "r") as genes_file:
        description_to_gene_mapping = json.load(genes_file)

    with open(gene_to_longest_sequence_file_path, "r") as genes_file:
        gene_to_longest_sequence = json.load(genes_file)

    columns_per_organism_count = 4
    general_columns_count = 4 + len(aa_list)
    aggregated_summary = defaultdict(list)
    for root, dirs, files in os.walk(results_directory):
        for file in files:
            if file == filename:
                sequence_name = Path(root).name.replace("-", "|")
                gene_name = description_to_gene_mapping.get(sequence_name)
                if gene_to_longest_sequence.get(gene_name) != sequence_name:
                    continue

                file_path = os.path.join(root, file)
                try:
                    workbook = openpyxl.load_workbook(filename=file_path)
                    worksheet = workbook.active
                except:
                    print(F"Failed to load: {file_path}")
                    continue

                original_headers = [cell.value for cell in worksheet[1]]
                headers = ["Sequence", "Gene"] + original_headers

                aggregated_summary["headers"] = headers
                optimization_method_to_scores = {}
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    if row[0] != optimization_method:
                        continue

                    organisms_optimization_index = "_".join(
                        F"{original_headers[i]}_{str(row[i])}" for i in
                        range(1, len(row) - general_columns_count, columns_per_organism_count)
                    )
                    key = "-".join([row[0], organisms_optimization_index])

                    # Needed to remove duplicates from the summary file
                    optimization_method_to_scores[key] = {
                        "sequence": sequence_name,
                        "gene": gene_name,
                        "data": row,
                    }

                for index, row in optimization_method_to_scores.items():
                    aggregated_summary[index].append(row)

    with open(F"{optimization_method}_summary.json", "w") as aggregated_summary_file:
        json.dump(aggregated_summary, aggregated_summary_file)


def convert_json_summary_to_xlsx(file_path: str) -> typing.Iterable[str]:
    with open(file_path, "r") as json_file:
        json_summary = json.load(json_file)

    with open("gene_to_longest_sequence.json", "r") as genes_file:
        gene_to_longest_sequence = json.load(genes_file)

    headers = json_summary.pop("headers")

    hit_genes = []
    missed_genes = set()
    for key, value in json_summary.items():
        summary_file_path = F"{key}-summary.xlsx"
        summary_workbook = openpyxl.Workbook()
        summary_worksheet = summary_workbook.active
        summary_worksheet.append(headers)

        for row_data in value:
            row = [row_data["sequence"], row_data["gene"]]
            row.extend(row_data["data"])
            summary_worksheet.append(row)
            hit_genes.append(row_data["gene"])

        summary_workbook.save(summary_file_path)

        missed_genes = set(gene_to_longest_sequence.keys()).difference(set(hit_genes))
    return missed_genes


def add_cell_with_value(worksheet, row: int, column: int, value: typing.Any) -> None:
    worksheet.insert_cols(column)
    worksheet.cell(row=row, column=column, value=value)


def initialize_column_headers(summary: typing.Dict, worksheet) -> typing.Sequence[str]:
    add_cell_with_value(worksheet=worksheet,
                        row=1,
                        column=worksheet.max_column,
                        value="optimization_method")

    add_cell_with_value(worksheet=worksheet,
                        row=1,
                        column=worksheet.max_column,
                        value="optimization_cub_index")

    organisms = summary["module_input"]["organisms"]
    # Display unwanted organisms first
    organisms.sort(key=lambda x: x.get("is_wanted"))
    for organism in organisms:
        organism_name = organism["name"]
        formatted_organism_name = "_".join(organism_name.split(" ")).lower()
        add_cell_with_value(worksheet=worksheet,
                            row=1,
                            column=worksheet.max_column,
                            value=F"{formatted_organism_name}_is_wanted")
        add_cell_with_value(worksheet=worksheet,
                            row=1,
                            column=worksheet.max_column,
                            value=F"{formatted_organism_name}_initial_gene_score")
        add_cell_with_value(worksheet=worksheet,
                            row=1,
                            column=worksheet.max_column,
                            value=F"{formatted_organism_name}_final_gene_score")
        add_cell_with_value(worksheet=worksheet,
                            row=1,
                            column=worksheet.max_column,
                            value=F"{formatted_organism_name}_dist_score")

    add_cell_with_value(worksheet=worksheet, row=1, column=worksheet.max_column, value="final_average_distance_score")
    add_cell_with_value(worksheet=worksheet, row=1, column=worksheet.max_column, value="final_weakest_link_score")
    add_cell_with_value(worksheet=worksheet, row=1, column=worksheet.max_column, value="final_ratio_score")

    for aa in aa_list:
        add_cell_with_value(worksheet=worksheet, row=1, column=worksheet.max_column, value=aa)

    add_cell_with_value(worksheet=worksheet, row=1, column=worksheet.max_column, value="number_of_iterations")
    add_cell_with_value(worksheet=worksheet, row=1, column=worksheet.max_column, value="run_time (seconds)")

    return [organism["name"] for organism in organisms]


def get_evaluation_summary(summary: typing.Dict[str, typing.Any]) -> typing.Dict[str, typing.Any]:
    final_evaluation = summary["final_evaluation"]
    for evaluation_summary in summary["evaluation"]:
        if evaluation_summary["average_distance_score"] == final_evaluation["average_distance_score"]:
            return evaluation_summary
    raise RuntimeError(F"Did not find a final evaluation for: {summary}")


def get_orf_summary(summary: typing.Dict[str, typing.Any]) -> typing.Dict[str, typing.Any]:
    if type(summary["orf"]) != list:
        return summary["orf"]

    final_evaluation = summary["final_evaluation"]
    for orf_summary in summary["orf"]:
        if orf_summary["final_sequence"] == final_evaluation["final_sequence"]:
            return orf_summary
    raise RuntimeError(F"Did not find an orf summary for: {summary}")


def update_from_summary(
        file_path: str,
        worksheet,
        ordered_organisms: typing.Optional[typing.Sequence[str]] = None,
) -> typing.Sequence[str]:
    with open(file_path, "r") as summary_file:
        summary = json.load(summary_file)

    if worksheet.max_row == 1:
        ordered_organisms = initialize_column_headers(summary=summary, worksheet=worksheet)

    optimization_method = summary["module_input"]["optimization_method"]
    optimization_cub_index = summary["module_input"]["optimization_cub_index"]
    formatted_optimization_cub_index = optimization_cub_index.lower()
    summary_row = [optimization_method, optimization_cub_index]

    evaluation_summary = get_evaluation_summary(summary)
    organisms = evaluation_summary["organisms"]
    for organism_name in ordered_organisms:
        organism = [matched_organism for matched_organism in organisms if
                    matched_organism["name"] == organism_name][0]
        summary_row.append(organism["is_wanted"])
        summary_row.append(organism[f"{formatted_optimization_cub_index}_initial_score"])
        summary_row.append(organism[f"{formatted_optimization_cub_index}_final_score"])
        summary_row.append(organism["dist_score"])

    summary_row.append(evaluation_summary["average_distance_score"])
    summary_row.append(evaluation_summary["weakest_link_score"])
    summary_row.append(evaluation_summary.get("ratio_score"))

    orf_summary = get_orf_summary(summary)
    aa_to_optimal_codon = orf_summary["aa_to_optimal_codon"]
    for aa in aa_list:
        summary_row.append(aa_to_optimal_codon.get(aa) or "")

    summary_row.append(orf_summary.get("iterations_count") or 1)
    summary_row.append(orf_summary["run_time"])

    worksheet.append(summary_row)

    return ordered_organisms


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Summary script parser")
    parser.add_argument('-m', '--method', type=str, required=True, help="Optimization method to collect results for.")

    args = parser.parse_args()
    requested_optimization_method = args.method
    # 1

    # generate_summary(results_directory=r"results\mcherry")
    generate_summary(results_directory=r"C:\projects\Igem_TAU_2021_moran\analysis\orf_model_analysis\results\endogenous\yfmJ-putative oxidoreductase\0.5")

    # 2
    root_dir = r"C:\projects\Igem_TAU_2021_moran\analysis\orf_model_analysis\results_human"
    # for file_name in os.listdir(root_dir):
    #     generate_summary(results_directory=os.path.join(root_dir, file_name))

    # 3
    # accumulate_summary_files(results_directory=root_dir,
    #                          description_to_gene_file_path="description_to_gene_mapping.json",
    #                          gene_to_longest_sequence_file_path="gene_to_longest_sequence.json",
    #                          optimization_method=requested_optimization_method)
    #
    # 4
    # accumulate_summary_files_to_json(results_directory=root_dir,
    #                                  description_to_gene_file_path="description_to_gene_mapping.json",
    #                                  gene_to_longest_sequence_file_path="gene_to_longest_sequence.json",
    #                                  optimization_method=requested_optimization_method)

    # 5
    # summary_files = [
    #     r"C:\Users\Kama\Documents\Moran\biomedical-engineering\microbiome-optimization\articles\ORF\orf_article_data\homo_sapiens\single_codon_diff_summary.json",
    #     r"C:\Users\Kama\Documents\Moran\biomedical-engineering\microbiome-optimization\articles\ORF\orf_article_data\homo_sapiens\single_codon_ratio_summary.json",
    #     r"C:\Users\Kama\Documents\Moran\biomedical-engineering\microbiome-optimization\articles\ORF\orf_article_data\homo_sapiens\zscore_bulk_aa_average_summary.json",
    #     r"C:\Users\Kama\Documents\Moran\biomedical-engineering\microbiome-optimization\articles\ORF\orf_article_data\homo_sapiens\zscore_single_aa_average_summary.json",
    # ]
    # missing_directories = set()
    # for file_path in summary_files:
    #     missing_directories.update(convert_json_summary_to_xlsx(file_path=file_path))
    #
    # with open("gene_to_longest_sequence.json", "r") as genes_file:
    #     gene_to_longest_sequence = json.load(genes_file)
    #
    # missed_genes = {gene: gene_to_longest_sequence[gene] for gene in missing_directories}
    # with open(F"missed_genes.json", "w") as missed_genes_file:
    #     json.dump(missed_genes, missed_genes_file)
